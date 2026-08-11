r"""etl.prefeitura.coleta_lixo — sync Betim household waste collection into
`coleta_lixo`.

Fonte: planilha XLSX oficial da Prefeitura de Betim, linkada a partir de
`https://www.betim.mg.gov.br/dias-e-horarios-da-coleta-de-lixo-domiciliar`
("Confira aqui os dias e horários em que o caminhão de lixo vai passar no
seu bairro"). Achada em sessão de descoberta 2026-08-11 — nenhuma das 6
cidades do projeto tinha fonte confirmada pra `coleta_lixo` até então.

    python -m etl.prefeitura.coleta_lixo
    python -m etl.prefeitura.coleta_lixo --id-municipio 3106705

═══ POR QUE XLSX, E NÃO A API DO GEOWEB ═══

Betim TEM uma API JSON sem auth pra isso — `GET
sistemas.betim.mg.gov.br/geoweb/coletaLixo/GetInfoColetaLixo?cep=X&num=Y`
(achada inspecionando `coletaLixo.entry.js`/`shared.entry.js`, o mesmo
GeoWeb que serve `infbasica/GetBairros`) — mas ela responde por ENDEREÇO
(CEP + número), não por bairro: dado um ponto, devolve o logradouro mais
próximo e a frequência/turno da coleta ali. Pra popular uma tabela por
BAIRRO a partir dela seria preciso escolher um endereço "representativo"
por bairro sem saber se o bairro inteiro tem o mesmo horário — o próprio
XLSX mostra que não tem sempre (ver armadilha 2). A planilha, em
contrapartida, JÁ é organizada por bairro (é o que a Prefeitura usa pra
publicar "dias e horários por bairro"), então é a fonte mais direta pro
schema desta tabela. A API de endereço fica registrada aqui pra quem um dia
quiser um recurso de "digite sua rua" mais preciso — está fora do escopo
desta rodada.

═══ AS ARMADILHAS, MEDIDAS AO VIVO (2026-08-11) ═══

1. **A URL DO ARQUIVO NÃO É ESTÁVEL** (`coleta_de_lixo_-_ok_03111757.xlsx` —
   o sufixo parece um upload id/timestamp). Por isso este módulo não fixa a
   URL do arquivo: busca a página HTML e extrai o `href` que termina em
   `.xlsx` a cada rodada (`_achar_link_xlsx`). Se a Prefeitura mudar o
   layout da página a ponto do link sumir, o módulo aborta alto em vez de
   silenciosamente coletar um XLSX antigo em cache.

2. **A PLANILHA TEM 40 ABAS: 39 "setor" (coleta comum) + "COLETA
   SELETIVA".** Cada aba de setor é um roteiro de caminhão, não uma lista de
   bairros — o MESMO bairro pode aparecer em dois setores com horários
   DIFERENTES (medido ao vivo, dado real gravado nesta rodada: "Icaivera"
   sai com `dias_semana=[terca,quinta,sabado]` sob `horario='A partir das
   07:00'` — setor 121, diurno — E com `dias_semana=[segunda,quinta]` sob
   `horario='A partir das 19:00'` — setor 008, noturno; são ruas diferentes
   do mesmo bairro, atendidas por rotas diferentes). Por isso a chave
   natural que este módulo usa é `(bairro, horário)`, não só `bairro`: duas
   linhas para o mesmo nome de bairro com horários diferentes SÃO duas
   linhas reais na tabela, refletindo a fonte. Dentro da MESMA chave, os
   dias de coleta de todos os setores que a compartilham são UNIDOS (união
   de conjuntos) — na única checagem cruzada feita nesta sessão (bairro
   "Centro", setores 013 e 030, ambos noturno/19:00) as duas ocorrências
   eram idênticas (todos os dias), então a união não perdeu nada; não há
   garantia formal de que isso vale pra todo par.

3. **AS CÉLULAS DE UMA LINHA DENTRO DE UM BLOCO NEM SEMPRE CONCORDAM ENTRE
   SI** — a aba `008` tem uma linha com `ICAIVERA` na segunda e quinta mas
   `ESPÍRITO SANTO` no sábado (a linha seguinte, a "dona" de fato do
   sábado, já lista `ESPÍRITO SANTO` nos três dias corretamente). Erro de
   preenchimento humano na planilha-fonte, não deste módulo. A escolha de
   parsing (armadilha 4) trata cada CÉLULA como uma observação independente
   `(bairro da célula, dia da coluna)`, não "o nome da primeira célula da
   linha vale pra linha inteira" — isso faz o erro se autocorrigir na
   agregação (Icaivera não ganha um sábado que não é dele; Espírito Santo já
   tinha o sábado certo na própria linha). Não filtra, nem around by
   comparação com nomes vizinhos; é uma consequência estrutural do jeito
   como a agregação funciona.

4. **NENHUMA VALIDAÇÃO DE ORTOGRAFIA.** A planilha tem o mesmo bairro
   grafado de formas diferentes em setores diferentes (`NITEROI` /
   `NTEROI`, `BANDEIRINHAS` / `BANDERINHAS`, `CACHOEIRA` / `CAHOEIRA`,
   `HOSPITAL` / `HOPITAL`). Este módulo NÃO tenta corrigir — normaliza só
   espaço e capitalização (`_normalizar_bairro`, `title()` preservando
   acentos) — porque "consertar" grafia sem confirmar qual é a oficial
   inventaria dado que a Prefeitura não escreveu. O efeito visível: alguns
   bairros aparecem como duas linhas quase-iguais na página. Registrado
   aqui pra quem quiser um dicionário de correção deliberado depois.

5. **NEM TODA CÉLULA É UM BAIRRO ADMINISTRATIVO** — a aba `COLETA SELETIVA`
   tem entradas como "Prefeitura De Betim", "Hospital Regional", "Colônia
   Hospital" e "Caçambas Ptb" (medido: as 4 caem na tabela gravada). São
   pontos de parada da rota (a própria Prefeitura, um hospital, um ponto de
   caçambas), do jeito que a fonte publica — este módulo não filtra por não
   ter como distinguir programaticamente "bairro" de "ponto de referência"
   sem uma lista oficial de bairros pra cruzar (que não existe pronta pro
   projeto ainda). Efeito na tela: a busca por bairro em `/coleta-lixo` vai
   achar essas ~4 linhas junto dos bairros de verdade.

6. **O TÍTULO DE ALGUMAS ABAS MENTE SOBRE O NÚMERO DO SETOR** (ex.: a aba
   `112` tem título "SETOR 113"; `101`/`102` têm título "SETOR 017"). O
   número do setor não é gravado em lugar nenhum do schema, então isso não
   afeta o dado gravado — só significa que este módulo lê o HORÁRIO do
   título (`_extrair_horario`, um regex `HH:MM`) e ignora o número do setor
   ali, usando o nome da ABA (não o título) só pra log/depuração.

═══ O QUE ESTE MÓDULO ESCREVE ═══

`coleta_lixo`, full-refresh por `(id_municipio, tipo)` via
`refresh_completo_seguro` — a tabela não tem constraint única (ninguém
upserta nela hoje) e o volume é pequeno (algumas centenas de linhas), então
apagar-e-reinserir por rodada é mais simples que manter uma chave e é seguro
contra regressão (o helper recusa apagar se a coleta nova vier menor que o
que já existe, a menos que `--permitir-reducao`).
"""
import argparse
import re
import sys
import unicodedata
from io import BytesIO

import openpyxl
import requests
from tenacity import retry, stop_after_attempt, wait_exponential

from etl.common import ID_MUNICIPIO_DEFAULT, get_supabase_client, refresh_completo_seguro

LOG = "[etl.prefeitura.coleta_lixo]"

# Só Betim: a fonte é a planilha da própria Prefeitura, não um fornecedor
# compartilhado. Ver `_conferir_cidade`.
ID_MUNICIPIO_BETIM = "3106705"

PAGINA_URL = "https://www.betim.mg.gov.br/dias-e-horarios-da-coleta-de-lixo-domiciliar"
TIMEOUT = 30

_SESSAO = requests.Session()
_SESSAO.headers["User-Agent"] = (
    "ControlePopular/1.0 (+https://github.com/FinweeJur/controle-popular)"
)

_RE_LINK_XLSX = re.compile(r'href="([^"]*\.xlsx[^"]*)"', re.IGNORECASE)
# Âncora em "PARTIR DAS" (não só em dígitos soltos): o título também carrega
# o número do setor ("SETOR 113 A partir das 07:00 hrs") e um regex de
# dígito solto pegaria o "113". Dois formatos de hora convivem na mesma
# planilha: abas de setor escrevem "07:00" (HH:MM), a aba de coleta
# seletiva escreve "7h" (H seguido de "h", sem minuto) — testado ao vivo
# 2026-08-11, os dois têm que casar.
_RE_HORARIO = re.compile(r"PARTIR\s+DAS\s+(\d{1,2})(?::(\d{2}))?\s*H?")

ABA_SELETIVA = "COLETA SELETIVA"

# Ordem canônica pra ordenar `dias_semana` na saída (mesmo vocabulário de
# `DIA_SEMANA_ICS` em `apps/web/lib/betim/servicos.ts`: sem acento,
# minúsculo).
_ORDEM_DIA = {
    "domingo": 0,
    "segunda": 1,
    "terca": 2,
    "quarta": 3,
    "quinta": 4,
    "sexta": 5,
    "sabado": 6,
}


def _sem_acento(texto: str) -> str:
    base = unicodedata.normalize("NFD", texto or "")
    return "".join(c for c in base if unicodedata.category(c) != "Mn")


@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=15))
def _get(url: str) -> requests.Response:
    resp = _SESSAO.get(url, timeout=TIMEOUT)
    resp.raise_for_status()
    return resp


def _achar_link_xlsx() -> str:
    """Busca a página da Prefeitura e extrai o `href` do XLSX vigente —
    ver armadilha 1 no topo do módulo sobre por que não é fixo."""
    resp = _get(PAGINA_URL)
    m = _RE_LINK_XLSX.search(resp.text)
    if not m:
        raise RuntimeError(
            f"{PAGINA_URL}: nenhum link .xlsx encontrado na página — "
            "a Prefeitura pode ter mudado o layout ou tirado o arquivo do ar."
        )
    link = m.group(1)
    if link.startswith("http"):
        return link
    return f"https://www.betim.mg.gov.br{link if link.startswith('/') else '/' + link}"


def _mapear_dia(texto) -> str | None:
    """Cabeçalho de coluna ("2ª FEIRA", "SÁBADO ", ...) -> dia canônico, ou
    `None` se a coluna não for um dia reconhecido (ex.: coluna vazia à
    direita da tabela)."""
    if not texto:
        return None
    t = _sem_acento(str(texto)).upper().strip()
    if t.startswith("2"):
        return "segunda"
    if t.startswith("3"):
        return "terca"
    if t.startswith("4"):
        return "quarta"
    if t.startswith("5"):
        return "quinta"
    if t.startswith("6"):
        return "sexta"
    if "SABADO" in t:
        return "sabado"
    if "DOMINGO" in t:
        return "domingo"
    return None


def _extrair_horario(titulo) -> str | None:
    texto = _sem_acento(str(titulo or "")).upper()
    m = _RE_HORARIO.search(texto)
    if not m:
        return None
    hora = m.group(1).zfill(2)
    minuto = (m.group(2) or "00").zfill(2)
    return f"A partir das {hora}:{minuto}"


def _normalizar_bairro(nome) -> str:
    limpo = " ".join(str(nome or "").split())
    return limpo.title()


def _e_titulo_de_bloco(valor) -> bool:
    """Uma linha "SETOR NNN ... A partir das HH:MM hrs" ou "A partir das
    Xh" abre um novo bloco (título + cabeçalho de dias na linha seguinte +
    linhas de bairro até a próxima linha em branco). Ver armadilha 2/3: cada
    aba pode ter mais de um bloco (`COLETA SELETIVA` tem dois)."""
    if not valor:
        return False
    t = _sem_acento(str(valor)).upper()
    return "PARTIR" in t or "SETOR" in t


def _parse_planilha(wb) -> dict[str, dict[tuple[str, str], set[str]]]:
    """Devolve `{tipo: {(bairro, horario): {dias}}}` pra todas as abas —
    `tipo` é "seletiva" só pra `COLETA SELETIVA`, "comum" pras 39 abas de
    setor. Ver armadilha 3 sobre por que a leitura é célula-a-célula."""
    por_tipo: dict[str, dict[tuple[str, str], set[str]]] = {"comum": {}, "seletiva": {}}

    for nome_aba in wb.sheetnames:
        tipo = "seletiva" if nome_aba == ABA_SELETIVA else "comum"
        ws = wb[nome_aba]
        linhas = [tuple(row) for row in ws.iter_rows(values_only=True)]
        n = len(linhas)
        i = 0
        blocos_na_aba = 0
        while i < n:
            titulo = linhas[i][0] if linhas[i] else None
            if not _e_titulo_de_bloco(titulo):
                i += 1
                continue
            horario = _extrair_horario(titulo)
            if horario is None:
                print(f"{LOG} AVISO: aba {nome_aba!r} linha {i+1}: título {titulo!r} sem HH:MM reconhecível — bloco pulado.")
                i += 1
                continue
            header = linhas[i + 1] if i + 1 < n else ()
            dias_colunas = [_mapear_dia(h) for h in header]
            blocos_na_aba += 1

            j = i + 2
            while j < n and any(v not in (None, "") for v in linhas[j]):
                for col, valor in enumerate(linhas[j]):
                    if col >= len(dias_colunas) or not dias_colunas[col]:
                        continue
                    if valor in (None, ""):
                        continue
                    bairro = _normalizar_bairro(valor)
                    chave = (bairro, horario)
                    por_tipo[tipo].setdefault(chave, set()).add(dias_colunas[col])
                j += 1
            i = j
        if blocos_na_aba == 0:
            print(f"{LOG} AVISO: aba {nome_aba!r} não tinha nenhum bloco reconhecível (título 'SETOR'/'A partir').")

    return por_tipo


def _linhas_para_gravar(id_municipio: str, agregado: dict[tuple[str, str], set[str]], tipo: str) -> list[dict]:
    linhas = []
    for (bairro, horario), dias in agregado.items():
        dias_ordenados = sorted(dias, key=lambda d: _ORDEM_DIA.get(d, 99))
        linhas.append(
            {
                "id_municipio": id_municipio,
                "bairro": bairro,
                "tipo": tipo,
                "dias_semana": dias_ordenados,
                "horario": horario,
            }
        )
    return linhas


def _conferir_cidade(id_municipio: str) -> None:
    if id_municipio != ID_MUNICIPIO_BETIM:
        raise RuntimeError(
            f"etl.prefeitura.coleta_lixo só sabe ler a planilha da Prefeitura de Betim "
            f"({ID_MUNICIPIO_BETIM}) — recebi id_municipio={id_municipio!r}. "
            "Rodar com outra cidade gravaria a agenda de Betim sob o id errado."
        )


def sync(id_municipio: str, *, permitir_reducao: bool = False) -> dict[str, int]:
    _conferir_cidade(id_municipio)

    link = _achar_link_xlsx()
    print(f"{LOG} planilha: {link}")
    resp = _get(link)
    wb = openpyxl.load_workbook(BytesIO(resp.content), data_only=True)
    print(f"{LOG} {len(wb.sheetnames)} aba(s) na planilha.")

    por_tipo = _parse_planilha(wb)

    client = get_supabase_client()
    resultado = {}
    for tipo, agregado in por_tipo.items():
        linhas = _linhas_para_gravar(id_municipio, agregado, tipo)
        if not linhas:
            print(f"{LOG} tipo={tipo}: nada extraído da planilha — não apago o que já existe.")
            resultado[tipo] = 0
            continue
        refresh_completo_seguro(
            client,
            "coleta_lixo",
            {"id_municipio": id_municipio, "tipo": tipo},
            linhas,
            permitir_reducao=permitir_reducao,
            rotulo="etl.prefeitura.coleta_lixo",
        )
        print(f"{LOG} tipo={tipo}: {len(linhas)} linha(s) (bairro, horário) gravada(s).")
        resultado[tipo] = len(linhas)
    return resultado


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--id-municipio", default=ID_MUNICIPIO_DEFAULT)
    parser.add_argument("--permitir-reducao", action="store_true")
    args = parser.parse_args()

    try:
        sync(args.id_municipio, permitir_reducao=args.permitir_reducao)
    except RuntimeError as e:
        print(f"{LOG} ABORT: {e}", file=sys.stderr)
        sys.exit(1)
