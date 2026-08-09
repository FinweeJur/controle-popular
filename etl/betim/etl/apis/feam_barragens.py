r"""etl.apis.feam_barragens — inventário anual de **barragens de MG** da FEAM
(Fundação Estadual do Meio Ambiente), por município.

Fonte: `https://feam.br/documents/d/feam/lista-de-barragens-2024-xlsx` — XLSX
público, sem chave e sem login, 249 barragens de mineração e indústria em MG.

POR QUE ESTA FONTE, TENDO JÁ O SNISB (`etl.apis.snisb_barragens`). O SNISB é
mais amplo (2.212 barragens em MG) mas o campo que responde "esta barragem está
perigosa AGORA" — `NIVEL_PERIGO` — vem **vazio em ~97% das linhas** lá. A FEAM
cobre menos barragens e responde justamente isso, preenchido:

  * `CONDIÇÃO DE ESTABILIDADE` (a DCE): Atestada 216 · Não Atestada 21 ·
    Não apresentou 10
  * `NÍVEL DE EMERGÊNCIA`: 0 → 231 · 1 → 11 · 2 → 4 · 3 → 3
  * `MÉTODO CONSTRUTIVO`: **Montante 34** — o método de Mariana e Brumadinho
  * `SUSPENSÃO`: Sim 32

São as três colunas que o eixo Ambiental queria do SIGIBAR e não pôde coletar
de lá (reCAPTCHA Enterprise, `docs/ambiental/F0-discovery.md` §13.1). Para as
barragens de MINERAÇÃO de MG, esta fonte entrega o mesmo conteúdo por um
caminho aberto.

═══ AS ARMADILHAS MEDIDAS AO VIVO (2026-08-09) ═══

1. **O CABEÇALHO ESTÁ NA LINHA 4, NÃO NA 1.** As três primeiras são um "Total:
   249", uma nota e uma linha em branco. Ler com `header=0` produz um DataFrame
   com colunas chamadas "Total" e "249".

2. **VARRER "LINHA NÃO TOTALMENTE VAZIA" DEVOLVE 251, NÃO 249.** Há duas linhas
   de rodapé em que as 20 colunas úteis estão vazias mas alguma célula fora do
   intervalo carrega formatação. O critério de linha válida aqui é ter **NOME DA
   BARRAGEM**, e com ele o total bate exatamente com o "Total: 249" que a
   própria planilha declara na célula A1 — conferência que este módulo faz
   (`_conferir_total`) e RECLAMA quando falha.

3. **COORDENADA VEM EM DOIS FORMATOS, E O SEGUNDO É SILENCIOSAMENTE ERRADO.**
   A maioria é string com sinal de grau grudado (`'-19.659893°'`), mas **5
   linhas trazem inteiro de 8 dígitos sem separador decimal** (`-19645284`, que
   é -19,645284). Gravar o inteiro cru numa coluna `numeric` põe a barragem a
   milhões de graus de distância, sem erro nenhum. Latitude e longitude podem
   ter tipos DIFERENTES na mesma linha (medido: `(-20431374, '-43.882828°')`).
   Depois de converter, a coordenada ainda é conferida contra a caixa de MG —
   fora dela vira `None`, porque coordenada errada num mapa é pior que
   coordenada ausente.

4. **`Id Sigibar` NÃO É CHAVE.** Duas barragens (Massa Falida da Mundo
   Mineração, em Rio Acima) trazem o literal **"Não cadastrado"** no lugar do
   número. A doc anterior do projeto afirmava "preenchido 249/249" — é 247
   numéricos + 2 textos iguais entre si. A chave natural aqui é
   `(município, nome da barragem)`, medida sem nenhuma colisão nas 249; o
   `id_sigibar` é guardado como TEXTO, para casar com o IDE quando existir e
   para não mentir quando não existir.

5. **O MESMO MÉTODO CONSTRUTIVO APARECE COM QUATRO GRAFIAS.** `Jusante` e
   `jusante`; `Linha de Centro` e `Linha de centro`; `Etapa única` com duas
   sequências de bytes distintas. Facetar pelo texto cru divide a mesma
   categoria em quatro baldes — e é justamente aqui que mora o `Montante`, que
   ninguém pode contar errado. Este módulo grava o rótulo canônico
   (`_canonizar_metodo`) e o texto original ao lado.

6. **CÉLULA "VAZIA" PODE SER `\xa0` (ESPAÇO NÃO SEPARÁVEL), NÃO `None`.**
   Quatro linhas têm isso em `CONDIÇÃO DE ESTABILIDADE`. `if valor:` diz que
   está preenchido, e o portal exibiria uma condição de estabilidade em branco
   como se fosse informação.

7. **ALTURA E VOLUME MISTURAM `str` E `int` NA MESMA COLUNA** (133 str × 114 int
   em ALTURA). Mesma classe do `_num` do CAP: converter por `Decimal(str(v))`.

8. **NÃO HÁ CÓDIGO IBGE.** Só `MUNICÍPIO` por nome (59 municípios distintos nas
   249 linhas). Mesma lacuna do SNISB, mesmo tratamento: casamento normalizado
   em código, e `--sondar` sem `--nome-municipio` lista o que a fonte conhece.

═══ FRESCOR: O [VERIFY] DA §5 ESTÁ RESOLVIDO ═══

A §5 do `F0-discovery.md` registrava que o dado é atualizado mensalmente num
painel cuja URL ninguém tinha achado. **Ela está dentro da própria planilha**,
como hyperlink da célula D2 ("Clique aqui"):

    https://app.powerbi.com/view?r=eyJrIjoiOThhNzgyMTQtNGU5Ny00Mzk0LWIzODItNDg3Nzk2MDlmYmEyIiwidCI6IjkyNGY5ODQ3LTI0MmUtNGE5YS04OTEzLTllNDM2NDliOWVhYSJ9

É um embed público do Power BI (sem login), e é um relatório DIFERENTE do que a
página `semad.mg.gov.br/painel-de-indicadores-do-sisema` divulga — os dois
tokens diferem. **Este módulo NÃO lê o painel**: dado de Power BI só sai por
API interna não documentada, enquanto o XLSX é estruturado e estável. O preço é
o frescor: o XLSX é anual (base 2024) e o painel é mensal. Quem precisar do
mensal tem a URL registrada aqui e em `PROVENIENCIA.json`.

═══ O QUE ESTE MÓDULO NÃO PROVA ═══

Cobertura é de barragem de MINERAÇÃO e INDÚSTRIA fiscalizada pela FEAM (209 +
40). Não cobre abastecimento de água, irrigação nem hidrelétrica — para essas,
`etl.apis.snisb_barragens`. "Zero barragens da FEAM" num município **não** é
"nenhuma barragem no município".

Três números diferentes seguem sem reconciliação (FEAM 249 × WFS IDE 259 ×
SNISB/ANM 320) — ver a nota na migration `0049`.

═══ O QUE ESTE MÓDULO ESCREVE ═══

`feam_barragens` — uma linha por (município, nome da barragem). Refresh total
filtrado por `id_municipio`, com o guarda de redução de
`refresh_completo_seguro`.

Uso:

    python -m etl.apis.feam_barragens --id-municipio 3106705 --sondar --nome-municipio Betim
    python -m etl.apis.feam_barragens --id-municipio 3106705 --sondar   # lista os municípios da fonte
    python -m etl.apis.feam_barragens --id-municipio 3106705
"""
import argparse
import io
import re
import sys
import unicodedata
from decimal import Decimal, InvalidOperation

import openpyxl
import requests

from etl.common import (
    ID_MUNICIPIO_DEFAULT,
    carregar_municipio,
    get_supabase_client,
    refresh_completo_seguro,
)

LOG = "[etl.apis.feam_barragens]"

XLSX_URL = "https://feam.br/documents/d/feam/lista-de-barragens-2024-xlsx"
TIMEOUT = 180
_UA = "ControlePopular/1.0 (+https://github.com/FinweeJur/controle-popular)"

# Armadilha 1: o cabeçalho real está na linha 4 (1-based); dado começa na 5.
LINHA_CABECALHO = 4
PRIMEIRA_LINHA_DADO = 5
N_COLUNAS = 20

# Índices 0-based dentro das 20 colunas úteis, na ordem da planilha.
(
    C_ITEM, C_SIGIBAR, C_EMPREENDEDOR, C_NOME, C_MUNICIPIO, C_URA, C_LAT, C_LON,
    C_ATIVIDADE, C_FINALIDADE, C_SITUACAO, C_ESTABILIDADE, C_METODO, C_ALTURA,
    C_VOLUME, C_RISCO, C_DANO, C_CLASSE, C_EMERGENCIA, C_SUSPENSAO,
) = range(N_COLUNAS)

# Armadilha 3: caixa de Minas Gerais, com folga. Coordenada fora disto é
# defeito de origem, não barragem em outro estado — a fonte é só de MG.
MG_LAT = (-23.5, -14.0)
MG_LON = (-51.5, -39.5)

# Armadilha 5: quatro grafias para as mesmas quatro categorias.
_METODOS_CANONICOS = ("Montante", "Jusante", "Linha de Centro", "Etapa única")


def _sessao() -> requests.Session:
    s = requests.Session()
    s.headers["User-Agent"] = _UA
    return s


def _normalizar(s) -> str:
    base = unicodedata.normalize("NFD", str(s or ""))
    sem_acento = "".join(c for c in base if unicodedata.category(c) != "Mn")
    # `\xa0` entra aqui como espaço e some no split (armadilha 6).
    return " ".join(sem_acento.upper().replace("\xa0", " ").split())


# ─────────────────────────── parsers de campo ──────────────────────────


def _txt(v) -> str | None:
    """Armadilha 6: `\xa0` não é conteúdo."""
    if v is None:
        return None
    s = str(v).replace("\xa0", " ").strip()
    return s or None


def _num(v) -> str | None:
    """Armadilha 7: a mesma coluna traz `str` e `int`."""
    s = _txt(v)
    if s is None:
        return None
    try:
        return str(Decimal(s))
    except InvalidOperation:
        return None


def _coord(v, faixa: tuple[float, float], rotulo: str) -> str | None:
    """Armadilha 3: string com `°` grudado OU inteiro de 8 dígitos sem
    separador decimal. Depois de converter, confere contra a caixa de MG —
    coordenada errada num mapa mente mais alto que coordenada ausente."""
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        d = Decimal(str(v))
        if abs(d) > 180:
            d = d / Decimal(10) ** 6
    else:
        s = re.sub(r"[^\d\.,\-]", "", str(v)).replace(",", ".")
        if not s:
            return None
        try:
            d = Decimal(s)
        except InvalidOperation:
            return None
        if abs(d) > 180:
            d = d / Decimal(10) ** 6
    if not (faixa[0] <= float(d) <= faixa[1]):
        print(f"{LOG} AVISO: {rotulo} {v!r} -> {d} fora da caixa de MG; gravando NULL.")
        return None
    return str(d)


def _canonizar_metodo(v) -> str | None:
    """Armadilha 5: `Jusante`/`jusante`, `Linha de Centro`/`Linha de centro`,
    e `Etapa única` com duas sequências de bytes. Sem isto, `Montante` — o
    método de Mariana e Brumadinho — pode ser contado a menos."""
    bruto = _txt(v)
    if bruto is None:
        return None
    alvo = _normalizar(bruto)
    for canonico in _METODOS_CANONICOS:
        if _normalizar(canonico) == alvo:
            return canonico
    return bruto


def _emergencia(v) -> int | None:
    s = _txt(v)
    if s is None:
        return None
    try:
        return int(Decimal(s))
    except InvalidOperation:
        return None


# ──────────────────────────── leitura do XLSX ──────────────────────────


def _baixar(sessao: requests.Session) -> bytes:
    r = sessao.get(XLSX_URL, timeout=TIMEOUT)
    r.raise_for_status()
    return r.content


def _linhas_brutas(conteudo: bytes) -> tuple[list[list], int | None]:
    """Devolve (linhas válidas, total declarado pela planilha em A1).

    Armadilha 2: o critério de linha válida é ter NOME DA BARRAGEM. Varrer
    "linha não totalmente vazia" devolve 251 para um inventário de 249."""
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    linhas: list[list] = []
    declarado: int | None = None
    for i, valores in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            # "Total" | "249"
            for cel in valores[:3]:
                try:
                    declarado = int(str(cel).strip())
                    break
                except (TypeError, ValueError):
                    continue
            continue
        if i < PRIMEIRA_LINHA_DADO:
            continue
        campos = list(valores[:N_COLUNAS]) + [None] * max(0, N_COLUNAS - len(valores))
        if _txt(campos[C_NOME]) is None:
            continue
        linhas.append(campos)
    wb.close()
    return linhas, declarado


def _conferir_total(linhas: list[list], declarado: int | None) -> None:
    if declarado is None:
        print(f"{LOG} AVISO: a planilha não declarou total em A1 — sem conferência.")
        return
    if len(linhas) != declarado:
        print(f"{LOG} AVISO: li {len(linhas)} barragem(ns) e a planilha declara "
              f"{declarado} em A1. Layout pode ter mudado — conferir antes de gravar.")
    else:
        print(f"{LOG} {len(linhas)} barragem(ns), igual ao total declarado pela planilha.")


def _parse(linha: list, id_municipio: str) -> dict:
    return {
        "id_municipio": id_municipio,
        "id_sigibar": _txt(linha[C_SIGIBAR]),   # armadilha 4: TEXTO, não número
        "nome": _txt(linha[C_NOME]),
        "empreendedor": _txt(linha[C_EMPREENDEDOR]),
        "ura": _txt(linha[C_URA]),
        "atividade": _txt(linha[C_ATIVIDADE]),
        "finalidade": _txt(linha[C_FINALIDADE]),
        "situacao": _txt(linha[C_SITUACAO]),
        "condicao_estabilidade": _txt(linha[C_ESTABILIDADE]),
        "metodo_construtivo": _canonizar_metodo(linha[C_METODO]),
        "metodo_construtivo_fonte": _txt(linha[C_METODO]),
        "altura_m": _num(linha[C_ALTURA]),
        "volume_reservatorio_m3": _num(linha[C_VOLUME]),
        "categoria_risco": _txt(linha[C_RISCO]),
        "dano_potencial": _txt(linha[C_DANO]),
        "classe": _txt(linha[C_CLASSE]),
        "nivel_emergencia": _emergencia(linha[C_EMERGENCIA]),
        "suspensao": _txt(linha[C_SUSPENSAO]),
        "latitude": _coord(linha[C_LAT], MG_LAT, "latitude"),
        "longitude": _coord(linha[C_LON], MG_LON, "longitude"),
        "municipio_fonte": _txt(linha[C_MUNICIPIO]),
    }


# ─────────────────────────────── coleta ────────────────────────────────


def coletar(id_municipio: str, nome_municipio: str) -> list[dict]:
    sessao = _sessao()
    linhas, declarado = _linhas_brutas(_baixar(sessao))
    _conferir_total(linhas, declarado)
    alvo = _normalizar(nome_municipio)
    return [
        _parse(l, id_municipio)
        for l in linhas
        if _normalizar(l[C_MUNICIPIO]) == alvo
    ]


# ─────────────────────────────── sondar ────────────────────────────────


def sondar(id_municipio: str, nome_municipio: str | None) -> None:
    """Sem gravar e sem ler `municipios` — funciona com a Neon fora do ar."""
    sessao = _sessao()
    linhas, declarado = _linhas_brutas(_baixar(sessao))
    _conferir_total(linhas, declarado)

    if not nome_municipio:
        contagem: dict[str, int] = {}
        for l in linhas:
            m = _txt(l[C_MUNICIPIO]) or "(sem município)"
            contagem[m] = contagem.get(m, 0) + 1
        print(f"{LOG} {len(contagem)} municípios na fonte (armadilha 8: casamento é por "
              f"NOME, confira o seu na lista) — top 15:")
        for m, n in sorted(contagem.items(), key=lambda kv: -kv[1])[:15]:
            print(f"       {m:<30} {n}")
        return

    achadas = [
        _parse(l, id_municipio) for l in linhas
        if _normalizar(l[C_MUNICIPIO]) == _normalizar(nome_municipio)
    ]
    print(f"\n{LOG} {nome_municipio}: {len(achadas)} barragem(ns)")
    for b in achadas:
        print(f"       {(b['nome'] or '(sem nome)'):<34} {b['empreendedor'] or '':<30}")
        print(f"         estabilidade={b['condicao_estabilidade']!r} "
              f"emergência={b['nivel_emergencia']!r} método={b['metodo_construtivo']!r} "
              f"risco={b['categoria_risco']!r}/dano={b['dano_potencial']!r} "
              f"situação={b['situacao']!r} suspensão={b['suspensao']!r}")


# ──────────────────────────────── sync ─────────────────────────────────


def sync(id_municipio: str, *, permitir_reducao: bool) -> None:
    cidade = carregar_municipio(id_municipio)
    if cidade["uf"] != "MG":
        raise RuntimeError(
            f"{LOG} {cidade['nome']}-{cidade['uf']} não é de Minas Gerais. O inventário "
            f"da FEAM é estadual — para barragem fora de MG use `etl.apis.snisb_barragens`."
        )
    print(f"{LOG} {cidade['nome']}-{cidade['uf']} ({id_municipio})")
    linhas = coletar(id_municipio, cidade["nome"])
    _gravar(cidade, linhas, permitir_reducao)


def _gravar(cidade: dict, linhas: list[dict], permitir_reducao: bool) -> None:
    if not linhas:
        # 59 municípios de 853 têm barragem da FEAM — o normal é não ter, e
        # refresh total com lista vazia apagaria histórico sem esta guarda.
        print(f"{LOG} nada coletado para {cidade['nome']} — NÃO apago o que já existe.")
        return
    client = get_supabase_client()
    refresh_completo_seguro(
        client,
        "feam_barragens",
        {"id_municipio": cidade["id_municipio"]},
        linhas,
        permitir_reducao=permitir_reducao,
        rotulo="etl.apis.feam_barragens",
    )
    print(f"{LOG} feam_barragens: {len(linhas)} linha(s) gravada(s).")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--id-municipio", default=ID_MUNICIPIO_DEFAULT)
    parser.add_argument("--permitir-reducao", action="store_true")
    parser.add_argument("--sondar", action="store_true", help="consulta e relata, NÃO grava, NÃO lê o banco")
    parser.add_argument("--nome-municipio", help="só com --sondar: a fonte não tem código IBGE")
    args = parser.parse_args()

    try:
        if args.sondar:
            sondar(args.id_municipio, args.nome_municipio)
        else:
            sync(args.id_municipio, permitir_reducao=args.permitir_reducao)
    except RuntimeError as e:
        print(f"{LOG} ABORT: {e}", file=sys.stderr)
        sys.exit(1)
