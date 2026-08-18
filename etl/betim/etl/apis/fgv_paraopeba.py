"""etl.apis.fgv_paraopeba — sync auditoria socioeconômica do Rio
Paraopeba (FGV) pra `paraopeba_saldo_municipio` e `paraopeba_iniciativas`.
Pedido do usuário 2026-07-24: Betim é um dos 26 municípios signatários do
Acordo Geral de Reparação (rompimento da barragem da Vale em
Brumadinho, 2019); a FGV audita a execução via
www18.fgv.br/projetorioparaopeba/acompanhamento-saldo-municipios.html.

Fonte real: a própria página tem uma seção "Dados abertos"
(projetorioparaopeba/dados-abertos.html) com planilha mensal --
`geral-{mes}-{ano}.xlsx` (aba "Iniciativas": uma linha por projeto, com
`Link Público da Iniciativa` e `Link para o termo de compromisso` reais
-- é isso que dá os "botões de acesso direto" pedidos) e
`financeiro-{ano}-{mes}.xlsx` (aba "Síntese Municípios Geral": saldo
financeiro agregado por município). Não precisa de scraping de HTML.

O nome do arquivo do mês mais recente nem sempre está disponível no dia
1 do mês seguinte (defasagem de publicação) -- este módulo tenta o mês
atual e recua até `MESES_TENTATIVAS` meses até achar um 200.

Régua da casa (medido em 15/08/2026): `www18.fgv.br/robots.txt` responde
`User-Agent: * / Disallow: /` -- o host inteiro pede para não ser
rastreado. Desde 17/08/2026 este módulo usa User-Agent honesto
(`ControlePopular/1.0 (+https://github.com/FinweeJur/controle-popular)`,
o mesmo dos coletores novos) e pausa de 1,5 s entre requisições, em vez
de fingir navegador. A aposentadoria deste ETL em favor de
`scripts/coletar-execucao-fgv.mts` (que cobre a bacia inteira) continua
sendo decisão do dono.
"""
import argparse
import io
import sys
import time
from datetime import date

import openpyxl
import requests

from etl.common import (
    ID_MUNICIPIO_DEFAULT,
    get_supabase_client,
    upsert_com_colunas_opcionais,
)

GERAL_URL = "https://www18.fgv.br/projetorioparaopeba/projetos-dados/dados-abertos/geral-{mes:02d}-{ano}.xlsx"
FINANCEIRO_URL = "https://www18.fgv.br/projetorioparaopeba/library/dados-abertos/financeiro-{ano}-{mes:02d}.xlsx"
MESES_TENTATIVAS = 4
NOME_MUNICIPIO = "Betim"
USER_AGENT = "ControlePopular/1.0 (+https://github.com/FinweeJur/controle-popular)"
PAUSA_SEGUNDOS = 1.5

COL_INICIATIVAS = {
    "id_fdi": "ID FDI",
    "titulo": "Título da Iniciativa",
    "municipios": "Municípios",
    "grupo": "Grupo de Iniciativas",
    "tipo_obrigacao": "Tipo de obrigação",
    "area_tematica": "Área Temática",
    "sub_area_tematica": "Sub Área Temática",
    "anexo": "Anexo",
    "status": "Status",
    "investimento": "Investimento",
    "percentual": "Percentual Realizado do Projeto",
    "valor_total": "Valor Total",
    "produtos_previstos": "Produtos previstos",
    "produtos_entregues": "Produtos entregues",
    "produtos_atraso": "Produtos em atraso",
    "equip_previstos": "Equipamentos Previstos",
    "equip_entregues": "Equipamentos Entregues",
    "link_publico": "Link Público da Iniciativa",
    "link_termo": "Link para o termo de compromisso",
}


def _baixar(url: str) -> bytes | None:
    headers = {"User-Agent": USER_AGENT}
    resp = requests.get(url, headers=headers, timeout=30)
    return resp.content if resp.status_code == 200 else None


def _achar_planilhas() -> tuple[bytes, bytes, str]:
    hoje = date.today()
    ano, mes = hoje.year, hoje.month
    for _ in range(MESES_TENTATIVAS):
        geral = _baixar(GERAL_URL.format(mes=mes, ano=ano))
        time.sleep(PAUSA_SEGUNDOS)  # uma requisição por vez, com pausa
        financeiro = _baixar(FINANCEIRO_URL.format(mes=mes, ano=ano))
        time.sleep(PAUSA_SEGUNDOS)
        if geral and financeiro:
            return geral, financeiro, f"{ano}-{mes:02d}"
        mes -= 1
        if mes == 0:
            mes = 12
            ano -= 1
    raise RuntimeError(
        f"nenhuma planilha encontrada nos últimos {MESES_TENTATIVAS} meses (última tentativa: {mes:02d}/{ano})"
    )


def _linhas_iniciativas(conteudo: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb["Iniciativas"]
    linhas = ws.iter_rows(values_only=True)
    cabecalho = next(linhas)
    return [dict(zip(cabecalho, linha)) for linha in linhas]


def _avanco_fisico(conteudo: bytes, nome_municipio: str) -> dict[str, dict]:
    """Aba "Avanço Físico": progresso FÍSICO real por projeto, filtrado pelo
    município. É de onde a FGV tira os % do "Avanço Físico" (Executado x
    Planejado) — diferente da coluna "Percentual Realizado do Projeto" da
    aba Iniciativas, que dá 0 pra vários projetos. Uma linha por (projeto,
    município). Retorna `{id_fdi: {"executado", "planejado"}}`."""
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb["Avanço Físico"]
    linhas = ws.iter_rows(values_only=True)
    cabecalho = next(linhas)
    mapa: dict[str, dict] = {}
    for linha in linhas:
        row = dict(zip(cabecalho, linha))
        if (row.get("Município") or "").strip() != nome_municipio:
            continue
        id_fdi = row.get("ID FDI")
        if id_fdi:
            mapa[str(id_fdi)] = {
                "executado": _num(row.get("Executado (%)")),
                "planejado": _num(row.get("Planejado (%)")),
            }
    return mapa


def _linha_saldo_municipio(conteudo: bytes, nome_municipio: str) -> dict | None:
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[1]]  # "Síntese Municípios Geral"
    linhas = ws.iter_rows(values_only=True)
    cabecalho = next(linhas)
    for linha in linhas:
        row = dict(zip(cabecalho, linha))
        municipio = (row.get("Município") or "").strip()
        if municipio == nome_municipio:
            return row
    return None


def _num(v) -> float | None:
    return float(v) if isinstance(v, (int, float)) else None


def sync(id_municipio: str) -> None:
    client = get_supabase_client()
    geral_bytes, financeiro_bytes, referencia = _achar_planilhas()
    print(f"[etl.apis.fgv_paraopeba] planilhas de referencia={referencia}")

    todas = _linhas_iniciativas(geral_bytes)
    betim = [r for r in todas if r.get(COL_INICIATIVAS["municipios"]) and NOME_MUNICIPIO in r[COL_INICIATIVAS["municipios"]]]
    avanco = _avanco_fisico(geral_bytes, NOME_MUNICIPIO)

    rows = []
    for r in betim:
        id_fdi = str(r.get(COL_INICIATIVAS["id_fdi"]))
        af = avanco.get(id_fdi, {})
        rows.append(
            {
                "id_municipio": id_municipio,
                "id_fdi": r.get(COL_INICIATIVAS["id_fdi"]),
                "titulo": r.get(COL_INICIATIVAS["titulo"]),
                "municipios_envolvidos": r.get(COL_INICIATIVAS["municipios"]),
                "grupo_iniciativas": r.get(COL_INICIATIVAS["grupo"]),
                "tipo_obrigacao": r.get(COL_INICIATIVAS["tipo_obrigacao"]),
                "area_tematica": r.get(COL_INICIATIVAS["area_tematica"]),
                "sub_area_tematica": r.get(COL_INICIATIVAS["sub_area_tematica"]),
                "anexo": r.get(COL_INICIATIVAS["anexo"]),
                "status": r.get(COL_INICIATIVAS["status"]),
                "investimento": _num(r.get(COL_INICIATIVAS["investimento"])),
                "valor_total": _num(r.get(COL_INICIATIVAS["valor_total"])),
                # `percentual_realizado` = Avanço Físico Executado (o % que a
                # FGV mostra), NÃO a coluna "Percentual Realizado do Projeto"
                # da aba Iniciativas (que dá 0 pra vários). `percentual_planejado`
                # = quanto deveria estar pronto (migration 0026, opcional).
                "percentual_realizado": af.get("executado"),
                "percentual_planejado": af.get("planejado"),
                "produtos_previstos": r.get(COL_INICIATIVAS["produtos_previstos"]),
                "produtos_entregues": r.get(COL_INICIATIVAS["produtos_entregues"]),
                "produtos_em_atraso": r.get(COL_INICIATIVAS["produtos_atraso"]),
                "equipamentos_previstos": r.get(COL_INICIATIVAS["equip_previstos"]),
                "equipamentos_entregues": r.get(COL_INICIATIVAS["equip_entregues"]),
                "link_publico": r.get(COL_INICIATIVAS["link_publico"]),
                "link_termo_compromisso": r.get(COL_INICIATIVAS["link_termo"]),
                "referencia": referencia,
            }
        )
    if rows:
        upsert_com_colunas_opcionais(
            client,
            "paraopeba_iniciativas",
            rows,
            colunas_opcionais=["percentual_planejado"],
            on_conflict="id_municipio,id_fdi",
        )
    print(f"[etl.apis.fgv_paraopeba] iniciativas_betim={len(rows)}")

    saldo = _linha_saldo_municipio(financeiro_bytes, NOME_MUNICIPIO)
    if saldo:
        client.table("paraopeba_saldo_municipio").upsert(
            {
                "id_municipio": id_municipio,
                "referencia": referencia,
                "valor_acordo_inicial": _num(saldo.get("Valor do Acordo Inicial (R$) (1)")),
                "valor_acordo_atual": _num(saldo.get("Valor do Acordo Atual (R$) (2) ")),
                "empenhos_autorizados": _num(saldo.get("Total de Empenhos Autorizados Atualizados (R$) (3)")),
                "saldo_teto": _num(saldo.get("Saldo Teto Considerando a Reserva (R$) (4)")),
            },
            on_conflict="id_municipio",
        ).execute()
        print(f"[etl.apis.fgv_paraopeba] saldo_municipio_gravado=sim")
    else:
        print(f"[etl.apis.fgv_paraopeba] AVISO: linha de saldo de {NOME_MUNICIPIO} nao encontrada")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--id-municipio", default=ID_MUNICIPIO_DEFAULT)
    args = parser.parse_args()
    try:
        sync(args.id_municipio)
    except RuntimeError as e:
        print(f"[etl.apis.fgv_paraopeba] ABORT: {e}", file=sys.stderr)
        sys.exit(1)
