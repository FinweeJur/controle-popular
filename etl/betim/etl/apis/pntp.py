"""etl.apis.pntp — sync PNTP (Programa Nacional de Transparência Pública,
ATRICON) evaluation for Betim's Prefeitura and Câmara into
`nota_transparencia`. Unlocks `/nota-betim`, previously "em breve" for lack
of a confirmed source.

Fonte: `https://radardatransparencia.atricon.org.br/dados/dados_pntp_{ano}.zip`
-- achado real 2026-07-23: **não precisa de scraping nem User-Agent
especial** (uma tentativa anterior tinha travado com ECONNRESET tentando
outra rota de acesso ao site; baixar o ZIP direto com `requests` funciona de
primeira). O ZIP contém `avaliacoes_pntp_{ano}.xlsx` (resumo por ente
avaliado, ~1.6MB) e `respostas/respostas_pntp_{ano}_{UF}.xlsx` (questionário
bruto por estado, 2-56MB cada) -- só o primeiro é necessário aqui.

Cobertura confirmada ao vivo: 853/853 municípios de MG têm avaliação de
Executivo E de Legislativo em 2025 -- dá pra calcular a posição no ranking
estadual sem precisar de fonte adicional nenhuma. A planilha é NACIONAL: o
recorte por estado sai da UF da cidade (ver `UF_POR_EXTENSO`), não de um
literal.
"""
import argparse
import io
import sys
import zipfile

import openpyxl
import requests

from etl.common import ID_MUNICIPIO_DEFAULT, carregar_municipio, get_supabase_client

ZIP_URL = "https://radardatransparencia.atricon.org.br/dados/dados_pntp_{ano}.zip"
PLANILHA_AVALIACOES = "pntp_{ano}/avaliacoes_pntp_{ano}.xlsx"

COL_UF = "UF"
COL_ESFERA = "Esfera"
COL_PODER = "Poder"
COL_IBGE = "ibge"
COL_INDICE = "Índice de Transparência"
COL_NIVEL = "Nível de Transparência"
COL_VAR_INDICE = "% de Variação de Índice"
COL_VAR_NIVEL = "Variação por Nível"
COL_HISTORICO_NIVEL = "Histórico do Nível"
COL_LINK = "Link Site Principal"


def _baixar_planilha(ano: int) -> list[dict]:
    resp = requests.get(ZIP_URL.format(ano=ano), timeout=180)
    resp.raise_for_status()
    z = zipfile.ZipFile(io.BytesIO(resp.content))
    dados = z.read(PLANILHA_AVALIACOES.format(ano=ano))
    wb = openpyxl.load_workbook(io.BytesIO(dados), read_only=True)
    ws = wb[wb.sheetnames[0]]
    linhas = ws.iter_rows(values_only=True)
    cabecalho = next(linhas)
    return [dict(zip(cabecalho, linha)) for linha in linhas]


# A planilha do PNTP escreve a UF por EXTENSO ("Minas Gerais"), e a tabela
# `municipios` guarda a sigla. Sem o de-para, o filtro não casa nada e o
# módulo termina com "registros=0" — que é indistinguível de "a cidade não
# foi avaliada".
UF_POR_EXTENSO = {
    "AC": "Acre", "AL": "Alagoas", "AP": "Amapá", "AM": "Amazonas",
    "BA": "Bahia", "CE": "Ceará", "DF": "Distrito Federal",
    "ES": "Espírito Santo", "GO": "Goiás", "MA": "Maranhão",
    "MT": "Mato Grosso", "MS": "Mato Grosso do Sul", "MG": "Minas Gerais",
    "PA": "Pará", "PB": "Paraíba", "PR": "Paraná", "PE": "Pernambuco",
    "PI": "Piauí", "RJ": "Rio de Janeiro", "RN": "Rio Grande do Norte",
    "RS": "Rio Grande do Sul", "RO": "Rondônia", "RR": "Roraima",
    "SC": "Santa Catarina", "SP": "São Paulo", "SE": "Sergipe",
    "TO": "Tocantins",
}


def _ranking_estadual(todas: list[dict], poder: str, uf_extenso: str) -> list[dict]:
    """Prefeituras (Executivo) ou Câmaras (Legislativo) municipais de UM
    estado, ordenadas por índice de transparência decrescente.

    A UF era o literal "Minas Gerais" DENTRO da função — não um default de
    argparse, então o guarda `conferir_defaults_de_cidade.py` não pegava.
    Rodar para São Paulo devolveu `registros=0` sem erro nenhum: o filtro
    simplesmente não casava, e o log dizia que a cidade não tinha avaliação.
    A planilha cobre o Brasil inteiro; era só o recorte que estava preso.
    """
    filtradas = [
        r
        for r in todas
        if r.get(COL_UF) == uf_extenso
        and r.get(COL_ESFERA) == "Municipal"
        and r.get(COL_PODER) == poder
    ]
    return sorted(filtradas, key=lambda r: r[COL_INDICE], reverse=True)


def _num_ou_nulo(valor):
    """Número, ou None quando a planilha escreve "não se aplica".

    O ATRICON grava **"-"** (não vazio, não zero) na variação de índice de
    quem não foi avaliado no ciclo anterior, e a coluna no banco é `numeric`.
    O resultado era `invalid input syntax for type numeric: "-"`, que derruba
    o INSERT INTEIRO — as duas linhas da cidade, Executivo e Legislativo.

    Ficou invisível por meses porque Betim, BH e São Paulo têm variação real.
    Só apareceu na primeira cidade PEQUENA, que estreou no ranking e portanto
    não tem com o que variar. E apareceu do pior jeito: o módulo já havia
    impresso índice, nível e posição na tela antes de morrer, então o log
    parecia um sucesso e a tabela ficava vazia.
    """
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return valor
    texto = str(valor).strip().replace("%", "").replace(",", ".")
    if texto in ("", "-", "--", "N/A", "n/a", "NA"):
        return None
    try:
        return float(texto)
    except ValueError:
        return None


def sync(id_municipio: str, ano: int) -> None:
    client = get_supabase_client()
    cidade = carregar_municipio(id_municipio)
    uf_extenso = UF_POR_EXTENSO.get(cidade["uf"])
    if not uf_extenso:
        raise RuntimeError(
            f"UF {cidade['uf']!r} desconhecida — complete UF_POR_EXTENSO. "
            "A planilha do PNTP escreve o estado por extenso."
        )

    todas = _baixar_planilha(ano)
    print(f"[etl.apis.pntp] avaliacoes_lidas={len(todas)} uf={uf_extenso}")

    rows = []
    for poder in ("Executivo", "Legislativo"):
        ranking = _ranking_estadual(todas, poder, uf_extenso)
        total = len(ranking)
        posicao = next(
            (i for i, r in enumerate(ranking, start=1) if r.get(COL_IBGE) == id_municipio),
            None,
        )
        entrada = next((r for r in ranking if r.get(COL_IBGE) == id_municipio), None)
        if not entrada:
            print(f"[etl.apis.pntp] AVISO: nenhuma avaliação de {poder} achada pra id_municipio={id_municipio}")
            continue
        rows.append(
            {
                "id_municipio": id_municipio,
                "ano": ano,
                "poder": poder,
                "indice_transparencia": entrada[COL_INDICE],
                "nivel_transparencia": entrada[COL_NIVEL],
                "variacao_indice": _num_ou_nulo(entrada.get(COL_VAR_INDICE)),
                "variacao_nivel": entrada.get(COL_VAR_NIVEL),
                "historico_nivel": entrada.get(COL_HISTORICO_NIVEL),
                "posicao_ranking_uf": posicao,
                "total_avaliados_uf": total,
                "link_site": entrada.get(COL_LINK),
            }
        )
        print(
            f"[etl.apis.pntp] {poder}: índice={entrada[COL_INDICE]:.4f} "
            f"nível={entrada[COL_NIVEL]} posição={posicao}/{total} em {cidade['uf']}"
        )

    if rows:
        client.table("nota_transparencia").upsert(rows, on_conflict="id_municipio,ano,poder").execute()
    print(f"[etl.apis.pntp] id_municipio={id_municipio} ano={ano} registros={len(rows)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--id-municipio", default=ID_MUNICIPIO_DEFAULT)
    parser.add_argument("--ano", type=int, default=2025)
    args = parser.parse_args()
    try:
        sync(args.id_municipio, args.ano)
    except RuntimeError as e:
        print(f"[etl.apis.pntp] ABORT: {e}", file=sys.stderr)
        sys.exit(1)
